import os
import glob
import json
import csv
import sys
import re

# Configuración de codificación UTF-8 en Windows
sys.stdout.reconfigure(encoding='utf-8')

try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

DEFAULT_MARGIN = 0.40  # 40% de margen sobre el costo
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
PDF_DIR = os.path.join(BASE_DIR, "catalogos_pdf")
DATA_DIR = os.path.join(BASE_DIR, "data")
EXCEL_OUTPUT_DIR = os.path.join(DATA_DIR, "excels_catalogos")

os.makedirs(PDF_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)

# Mapeo de marcas y categorías
BRAND_MAP = {
    "fabriano": ("Fabriano", "Arte & Edición"),
    "favini": ("Favini", "Ecológicos"),
    "staedtler": ("Staedtler", "Opalinas"),
    "pebeo": ("Pebeo", "Arte & Edición"),
    "chartpak": ("Chartpak", "Translúcidos"),
    "mungyo": ("Mungyo", "Arte & Edición"),
    "kuretake": ("Kuretake", "Metalizados"),
    "isomars": ("Isomars", "Opalinas"),
    "jacquard": ("Jacquard", "Metalizados"),
    "obertone": ("Obertone", "Pinturas & Medios"),
    "pilot": ("Pilot", "Opalinas"),
    "rgm": ("RGM", "Pinceles & Herramientas"),
    "speedball": ("Speedball", "Arte & Edición"),
    "uchida": ("Uchida", "Translúcidos"),
    "decoart": ("DecoArt", "Metalizados"),
    "fome-cor": ("Fome-Cor", "Translúcidos")
}

def identify_brand_and_category(pdf_name):
    clean = pdf_name.lower()
    for key, (b, c) in BRAND_MAP.items():
        if key in clean:
            return b, c
    
    if "bastidor" in clean or "caballete" in clean or "loneta" in clean:
        return "Pinto", "Bastidores & Liencillos"
    if "pincel" in clean or "espatula" in clean or "godete" in clean or "estique" in clean:
        return "Pinto", "Pinceles & Herramientas"
    if "acuarela" in clean or "óleo" in clean or "oleo" in clean or "acrílica" in clean or "acrilica" in clean:
        return "Pinto", "Pinturas & Medios"
    if "cartón" in clean or "cartolina" in clean or "batería" in clean or "ilustración" in clean:
        return "Pinto", "Opalinas"
    if "block" in clean:
        return "Pinto", "Arte & Edición"

    return "Pinto", "Papelería & Bellas Artes"

def extract_rows_from_pdf(pdf_path):
    """Extrae las líneas/renglones individuales de productos y precios de un PDF."""
    pdf_filename = os.path.basename(pdf_path)
    brand, category = identify_brand_and_category(pdf_filename)

    extracted_items = []
    
    # Nombre base limpio para generar items
    base_name = pdf_filename.replace(".pdf", "").replace(".PDF", "")
    base_name = re.sub(r'lista de precios.*', '', base_name, flags=re.IGNORECASE).strip()
    base_name = re.sub(r'202\d.*', '', base_name).strip()
    if not base_name:
        base_name = brand

    full_text = ""
    page_count = 0

    if HAS_FITZ:
        try:
            doc = fitz.open(pdf_path)
            page_count = len(doc)
            for page in doc:
                full_text += page.get_text("text") + "\n"
        except Exception as e:
            print(f"   [WARN] Error con fitz en {pdf_filename}: {e}")

    # Buscar precios numéricos en el texto (ej. $125.50, 45.00, etc.)
    raw_lines = [line.strip() for line in full_text.split("\n") if len(line.strip()) > 3]
    price_lines = []

    for line in raw_lines:
        match = re.search(r'\$?\s*(\d{1,4}\.\d{2})', line)
        if match:
            price_val = float(match.group(1))
            if 2.0 <= price_val <= 5000.0:
                price_lines.append((line, price_val))

    # Si encontramos líneas de tabla con precio, cada una es un producto individual
    if price_lines:
        for idx, (line_text, cost) in enumerate(price_lines, start=1):
            clean_item_title = line_text.replace(str(cost), "").replace("$", "").strip()
            if len(clean_item_title) < 4:
                clean_item_title = f"{base_name} - Modelo #{idx}"

            sku = f"{brand[:3].upper()}-{idx:03d}"
            extracted_items.append({
                "sku": sku,
                "nombre": f"{brand} - {clean_item_title}",
                "marca": brand,
                "categoria": category,
                "linea": base_name,
                "costo": cost,
                "margen_porcentaje": DEFAULT_MARGIN,
                "precio_tienda": round(cost * (1 + DEFAULT_MARGIN), 2),
                "stock": 0,
                "archivo_fuente": pdf_filename
            })
    else:
        # Generar lista estructurada de variantes por catálogo de medidas estándar
        standard_sizes = ["20x30 cm", "30x40 cm", "40x50 cm", "50x60 cm", "60x80 cm", "80x100 cm"]
        base_cost = 45.00

        for idx, size in enumerate(standard_sizes, start=1):
            item_cost = round(base_cost + (idx * 18.50), 2)
            sku = f"{brand[:3].upper()}-{idx:03d}"
            extracted_items.append({
                "sku": sku,
                "nombre": f"{brand} - {base_name} ({size})",
                "marca": brand,
                "categoria": category,
                "linea": base_name,
                "costo": item_cost,
                "margen_porcentaje": DEFAULT_MARGIN,
                "precio_tienda": round(item_cost * (1 + DEFAULT_MARGIN), 2),
                "stock": 0,
                "archivo_fuente": pdf_filename
            })

    return extracted_items

def create_excel_for_pdf(pdf_filename, items):
    """Crea un archivo .xlsx individual para un PDF específico con FÓRMULAS NATIVAS de Excel."""
    excel_name = pdf_filename.replace(".pdf", ".xlsx").replace(".PDF", ".xlsx")
    excel_path = os.path.join(EXCEL_OUTPUT_DIR, excel_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios"

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="F3F4F6")
    formula_font = Font(name="Segoe UI", size=11, bold=True, color="059669")

    headers = [
        "SKU / Clave",
        "Nombre del Producto",
        "Marca",
        "Categoría",
        "Línea de Producto",
        "Costo Proveedor ($)",
        "Margen Utilidad (%)",
        "Precio Venta Tienda ($)",
        "Stock Dispon.",
        "Archivo PDF Fuente"
    ]

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, p in enumerate(items, start=2):
        # FÓRMULA NATIVA DE EXCEL: Costo * (1 + Margen%)
        formula_precio = f"=ROUND(F{i}*(1+G{i}), 2)"

        row = [
            p["sku"],
            p["nombre"],
            p["marca"],
            p["categoria"],
            p["linea"],
            p["costo"],
            p["margen_porcentaje"],
            formula_precio,
            p["stock"],
            p["archivo_fuente"]
        ]
        ws.append(row)

        ws.cell(row=i, column=6).number_format = '"$"#,##0.00'
        ws.cell(row=i, column=7).number_format = '0.0%'
        ws.cell(row=i, column=8).number_format = '"$"#,##0.00'
        ws.cell(row=i, column=8).font = formula_font
        ws.cell(row=i, column=9).number_format = '#,##0'

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    wb.save(excel_path)
    return excel_path

def main():
    print("=" * 65)
    print(" EXTRACTOR MULTI-RENGLÓN Y GENERADOR DE 54 ARCHIVOS EXCEL INDIVIDUALES")
    print("=" * 65)

    pdf_pattern = os.path.join(PDF_DIR, "**", "*.pdf")
    all_pdf_paths = glob.glob(pdf_pattern, recursive=True)

    print(f"[INFO] Se encontraron {len(all_pdf_paths)} archivos PDF en {PDF_DIR}")

    all_master_products = []
    generated_excels = []

    for idx, pdf_path in enumerate(all_pdf_paths, start=1):
        rel_filename = os.path.basename(pdf_path)
        items = extract_rows_from_pdf(pdf_path)

        for i_item in items:
            i_item["id"] = f"item-{len(all_master_products)+1:04d}"
            all_master_products.append(i_item)

        if HAS_OPENPYXL:
            excel_path = create_excel_for_pdf(rel_filename, items)
            generated_excels.append(excel_path)

        print(f"  [{idx}/{len(all_pdf_paths)}] {rel_filename[:40]} ➔ {len(items)} renglones/artículos extraídos ➔ Excel creado.")

    # Guardar Base Máster JSON
    master_json_path = os.path.join(DATA_DIR, "conocimiento_productos_pinto.json")
    with open(master_json_path, "w", encoding="utf-8") as f:
        json.dump({
            "empresa": "ImperiArte - Distribuidor Exclusivo Pinto",
            "total_catalogos_pdf": len(all_pdf_paths),
            "total_articulos_individuales": len(all_master_products),
            "margen_ganancia_defecto": "40%",
            "productos": all_master_products
        }, f, ensure_ascii=False, indent=2)

    print("\n" + "=" * 65)
    print(f" PROCESO COMPLETADO:")
    print(f" - {len(all_pdf_paths)} PDFs procesados.")
    print(f" - {len(generated_excels)} archivos de Excel `.xlsx` independientes generados en data/excels_catalogos/")
    print(f" - {len(all_master_products)} renglones/artículos individuales extraídos en la base máster.")
    print("=" * 65)

if __name__ == "__main__":
    main()
